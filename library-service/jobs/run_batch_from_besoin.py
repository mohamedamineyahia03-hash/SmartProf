"""Batch generation driven by BESOIN.xlsx (the user's exact exercise-count
need per level/subject/domain) -- built 2026-09-04 alongside the no-human-
review pipeline change (generation/review.py, generation/publish.py).

For each (niveau, matiere, domaine, count) row, generates `count` published
exercises for that domain, cycling across its skills/formats for variety.
Each attempt goes through generate -> validate -> automated Opus review ->
publish, with bounded regeneration on rejection (jobs/run_crawl.py's
generate_validate_publish, reused here so both entry points share the exact
same quality gate).

No crawler/web source needed -- BESOIN.xlsx rows are curriculum-derived
(the official Tunisian programme, already the Main App's own domain/skill
labels), not scraped content, so each domain gets one lightweight in-house
Source row (license_status="explicit_open", same pattern as
seed/seed_expression_recitation.py) instead of a crawled URL. Still
generation_run_id-backed like every other exercise -- no shortcut around
the "always generated, never copied" schema rule.

Usage: python jobs/run_batch_from_besoin.py <path-to-BESOIN.xlsx> [sheet_name]
Defaults to the LAST sheet in the workbook (BESOIN.xlsx's "Feuil2" is the
settled need, 6216 total -- "Feuil1" was an earlier draft with some rows at 0).

Requires ANTHROPIC_API_KEY (real generation + review) and the Main App
server running at MAIN_APP_URL (curriculum-schema lookups over HTTP -- see
curriculum_client.py; library-service never reads the Main App's DB
directly). Without a key, runs in the existing dry-run mode (placeholder
content, review auto-passes) so the whole pipeline stays testable.
"""

import os
import sys
import unicodedata

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # library-service/

import openpyxl  # noqa: E402

from app import app  # noqa: E402
from db import db  # noqa: E402
from curriculum_client import fetch_curriculum_schema  # noqa: E402
from jobs.run_crawl import generate_validate_publish  # noqa: E402
from models import Source  # noqa: E402

LEVEL_NAME_TO_CODE = {
    "1ere annee": "1", "2eme annee": "2", "3eme annee": "3",
    "4eme annee": "4", "5eme annee": "5", "6eme annee": "6",
}
SUBJECT_NAME_TO_CODE = {
    "arabe": "ar", "francais": "fr", "anglais": "en",
    "mathematiques": "math", "eveil scientifique": "science",
}

# Safety cap per domain: stop after this many total attempts even if the
# count target isn't reached, so one persistently broken domain (e.g. no
# curriculum match, or every generation getting rejected) can't loop
# forever / burn unbounded API spend.
MAX_ATTEMPTS_PER_UNIT = 3


def _norm(text):
    text = unicodedata.normalize("NFKD", text or "").encode("ascii", "ignore").decode("ascii")
    return text.strip().lower()


def load_needs(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[-1]]
    needs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        niveau, matiere, domaine, count = row[0], row[1], row[2], row[3]
        if not count:
            continue
        level_code = LEVEL_NAME_TO_CODE.get(_norm(niveau))
        subject_code = SUBJECT_NAME_TO_CODE.get(_norm(matiere))
        if not level_code or not subject_code:
            print(f"  ! unrecognized niveau/matiere: {niveau!r} / {matiere!r} -- skipped")
            continue
        needs.append((level_code, subject_code, domaine, int(count)))
    return needs


def _find_domain(level_code, subject_code, domaine_label):
    """Matches a BESOIN.xlsx domain label to the Main App's curriculum
    domain by name_fr (case/accent-insensitive) -- both were authored from
    the same official programme file, so names should align; a miss means
    the curriculum doesn't have this section yet (see the T1 gap audit)."""
    schema = fetch_curriculum_schema(level=level_code, subject=subject_code)
    target = _norm(domaine_label)
    for domain in schema:
        if _norm(domain["name_fr"]) == target:
            return domain
    return None


def _get_or_create_source(level_code, subject_code, domain):
    """One lightweight in-house Source per domain, reused across every
    exercise generated for it -- see module docstring."""
    url = f"local://besoin_batch_v1#{level_code}/{subject_code}/{domain['domain']}"
    existing = Source.query.filter_by(url=url).first()
    if existing:
        return existing
    source = Source(
        url=url,
        title=f"Programme officiel -- {domain['name_fr']}",
        license_status="explicit_open",
        subject_code=subject_code,
        level_code=level_code,
        domain_hint=domain["domain"],
        trimester_hint="T1",
        region_scope="tunisia_official",
        content_snapshot=(
            f"Domaine du programme officiel tunisien : {domain['name_fr']} "
            f"({domain['name_ar']}), niveau {level_code}."
        ),
        status="used_for_generation",
    )
    db.session.add(source)
    db.session.commit()
    return source


def run_domain(level_code, subject_code, domaine_label, count):
    """Generates up to `count` published exercises for one BESOIN.xlsx row.
    Returns (published, permanently_rejected)."""
    domain = _find_domain(level_code, subject_code, domaine_label)
    if domain is None:
        print(f"! {level_code}/{subject_code}/{domaine_label}: no matching curriculum domain -- skipped ({count} needed)")
        return 0, 0
    if not domain["skills"]:
        print(f"! {level_code}/{subject_code}/{domaine_label}: domain has no skills defined -- skipped ({count} needed)")
        return 0, 0

    source = _get_or_create_source(level_code, subject_code, domain)
    published = rejected = 0
    attempts = 0
    max_attempts = count * MAX_ATTEMPTS_PER_UNIT

    while published < count and attempts < max_attempts:
        skill = domain["skills"][attempts % len(domain["skills"])]
        formats = skill["exercise_formats"] or ["qcm"]
        exercise_format = formats[attempts % len(formats)]
        attempts += 1

        outcome, _exercise = generate_validate_publish(source, level_code, subject_code, domain, skill, exercise_format)
        if outcome == "published":
            published += 1
        elif outcome == "rejected":
            rejected += 1
        # "error" (API failure) doesn't count either way -- just move to the next slot.

    if published < count:
        print(f"  ! stopped after {attempts} attempts, {published}/{count} published for {domain['domain']}")
    return published, rejected


def run_batch(xlsx_path, sheet_name=None):
    needs = load_needs(xlsx_path, sheet_name)
    total_needed = sum(n[3] for n in needs)
    print(f"{len(needs)} domain(s), {total_needed} exercise(s) needed in total.\n")

    grand_published = grand_rejected = 0
    for level_code, subject_code, domaine_label, count in needs:
        published, rejected = run_domain(level_code, subject_code, domaine_label, count)
        grand_published += published
        grand_rejected += rejected
        print(f"{level_code}/{subject_code}/{domaine_label}: {published}/{count} published, {rejected} permanently rejected")

    print(f"\nDone: {grand_published}/{total_needed} published, {grand_rejected} permanently rejected.")
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("(ran in dry-run mode -- set ANTHROPIC_API_KEY for real generation)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python jobs/run_batch_from_besoin.py <path-to-BESOIN.xlsx> [sheet_name]")
        sys.exit(1)
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    with app.app_context():
        db.create_all()
        run_batch(sys.argv[1], sheet)
